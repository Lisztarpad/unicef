import streamlit as st
import pandas as pd
from datetime import datetime
import io
import re
from openpyxl.utils import get_column_letter

# 注入全局 CSS 样式
st.markdown(
    """
    <style>
        .block-container {
            padding-top: 1rem !important;
            padding-bottom: 4rem !important;
        }
    </style>
    """,
    unsafe_allow_html=True
)

# Set page configuration
st.set_page_config(page_title="RAS Anomaly Cases Detector", layout="wide")

st.title("RAS Anomaly Cases Detector")
st.write("Please upload the required reports to find anomaly cases and actions-taking progress.")
st.write("")

# --- Initialize Session State for Duplicate Confirmation ---
if 'dup_confirmed' not in st.session_state:
    st.session_state.dup_confirmed = False


def reset_dup_state():
    """Reset the confirmation state when new files are uploaded."""
    st.session_state.dup_confirmed = False


# =========================================================
# CONFIGURATION: Status Mapping Dictionary
# =========================================================
VALID_STATUS_MAPPING = {
    'Pending RAS discussion': 'Draft',
    'Awaiting Approval': 'Pending Requisition Approval',
    'Pending RAS Validation': 'Pending Requisition Approval',
    'Longlist QA Manual Check': 'Longlist',
    'Technical Assessment Creation': 'Technical Assessment Phase',
    'Upload online assessment': 'Technical Assessment Phase',
    'Interview details setup': 'Interview Phase',
    'Interview Scheduling': 'Interview Phase',
    'Interviewing': 'Interview Phase',
    'Longlisting': 'Longlist',
    'Processing': 'Longlist'
}


# --- Helper Functions ---
def load_csv(uploaded_file):
    try:
        return pd.read_csv(uploaded_file, encoding='utf-8')
    except UnicodeDecodeError:
        uploaded_file.seek(0)
        return pd.read_csv(uploaded_file, encoding='latin1')


def color_cells(row):
    styles = [''] * len(row)
    red_style = 'background-color: #ffcccc; color: #cc0000; font-weight: bold;'
    if row['Issue Type'] == 'Status Mismatch':
        styles[row.index.get_loc('Intella State')] = red_style
        styles[row.index.get_loc('TMS Status')] = red_style
    elif row['Issue Type'] == 'Offer Issued & Inactive > 60 Days':
        styles[row.index.get_loc('Last Updated Date')] = red_style
        styles[row.index.get_loc('Days Inactive')] = red_style
    return styles


def style_summary_cells(row):
    styles = [''] * len(row)
    gray_style = 'background-color: #D3D3D3; font-weight: bold;'
    if row['Issue Type'] == 'Total Anomalies':
        styles[row.index.get_loc('Responsible Person')] = gray_style
        styles[row.index.get_loc('Issue Type')] = gray_style
        styles[row.index.get_loc('Case Count')] = gray_style
    return styles


# =========================================================
# APP LAYOUT: Tabs
# =========================================================
tab1, tab2 = st.tabs(["📊 Generate Reports", "📈 Track Progress"])

# ---------------------------------------------------------
# TAB 1: GENERATE REPORTS (Data Processing)
# ---------------------------------------------------------
with tab1:
    st.write("Please upload the required reports to find anomaly cases and their responsible persons.")

    # File uploaders
    col1, col2, col3 = st.columns(3)
    with col1:
        file_intella = st.file_uploader("1. Intella Report (.csv)", type=['csv'], on_change=reset_dup_state)
    with col2:
        file_tms_reg = st.file_uploader("2. TMS Regular Report (.csv)", type=['csv'], on_change=reset_dup_state)
    with col3:
        file_tms_ffi = st.file_uploader("3. TMS FFI Report (.csv)", type=['csv'], on_change=reset_dup_state)

    if file_intella and file_tms_reg and file_tms_ffi:
        try:
            df_intella = load_csv(file_intella)
            df_tms_reg = load_csv(file_tms_reg)
            df_tms_ffi = load_csv(file_tms_ffi)

            tms_cols = ['Requisition Number', 'Requisition status']

            if not all(col in df_tms_reg.columns for col in tms_cols) or not all(
                    col in df_tms_ffi.columns for col in tms_cols):
                st.error(
                    "Error: Both TMS Regular and TMS FFI reports must contain 'Requisition Number' and 'Requisition status' columns.")
            elif 'job_id' not in df_intella.columns:
                st.error("Error: Intella Report must contain a 'job_id' column.")
            else:
                # 1. Clean Intella Report
                df_intella = df_intella.dropna(subset=['job_id'])
                df_intella['job_id'] = df_intella['job_id'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
                df_intella = df_intella[df_intella['job_id'] != '']

                # 2. Extract specific columns and Combine TMS files
                df_tms_reg_sub = df_tms_reg[tms_cols].copy()
                df_tms_ffi_sub = df_tms_ffi[tms_cols].copy()
                df_tms = pd.concat([df_tms_reg_sub, df_tms_ffi_sub], ignore_index=True)

                df_tms['Requisition Number'] = df_tms['Requisition Number'].astype(str).str.replace(r'\.0$', '',
                                                                                                    regex=True).str.strip()
                df_tms = df_tms[df_tms['Requisition Number'] != '']

                # 3. Duplicate Checking and User Confirmation
                dup_count = df_tms.duplicated(subset=['Requisition Number']).sum()

                if dup_count > 0:
                    if not st.session_state.dup_confirmed:
                        st.warning(f"⚠️ Found {dup_count} duplicated Requisition Numbers in the combined TMS reports.")
                        show_dups = df_tms[df_tms.duplicated(subset=['Requisition Number'], keep=False)].sort_values(
                            by='Requisition Number')
                        st.dataframe(show_dups)

                        if st.button("✅ Confirm and Remove Duplicates"):
                            st.session_state.dup_confirmed = True
                            st.rerun()
                        st.stop()
                    else:
                        st.success(f"✅ User confirmed: {dup_count} duplicated records have been removed.")
                        df_tms = df_tms.drop_duplicates(subset=['Requisition Number'])
                else:
                    st.info("✅ No duplicated Requisition Numbers found in the combined TMS reports.")
                    df_tms = df_tms.drop_duplicates(subset=['Requisition Number'])

                # 4. Merge Data (Generate MR)
                with st.spinner('Generating Merged Report and processing anomalies...'):
                    mr = pd.merge(df_intella, df_tms, left_on='job_id', right_on='Requisition Number', how='left')

                    # 恢复：Export MR Module
                    st.markdown("---")
                    st.subheader("📄 Merged Report (MR)")
                    st.write(f"The combined dataset contains {len(mr)} rows.")

                    output_mr = io.BytesIO()
                    with pd.ExcelWriter(output_mr, engine='openpyxl') as writer:
                        mr.to_excel(writer, index=False, sheet_name='MR')
                        ws_mr = writer.sheets['MR']
                        for idx, col in enumerate(mr.columns):
                            ws_mr.column_dimensions[get_column_letter(idx + 1)].width = 15

                    st.download_button(
                        label="📥 Download Merged Report (Excel)",
                        data=output_mr.getvalue(),
                        file_name='Merged_Report_MR.xlsx',
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    )

                    required_cols = ['state', 'number', 'assigned_to', 'sys_updated_on']
                    missing_cols = [col for col in required_cols if col not in mr.columns]

                    if missing_cols:
                        st.error(f"Error: Missing columns in Intella data: {', '.join(missing_cols)}")
                    else:
                        mr['sys_updated_on_dt'] = pd.to_datetime(mr['sys_updated_on'], errors='coerce', dayfirst=True)
                        today = pd.Timestamp.today()

                        ras_anomalies = []
                        non_ras_anomalies = []
                        not_in_tms_cases = []

                        # 5. Find Anomaly Cases
                        for index, row in mr.iterrows():
                            state = str(row['state']).strip() if pd.notna(row['state']) else ""
                            case_num = str(row['number'])
                            assigned_to = str(row['assigned_to'])
                            tms_num = str(row['job_id']).strip() if pd.notna(row['job_id']) else ""

                            if 'assignment_group' in mr.columns and pd.notna(row['assignment_group']):
                                assign_group = str(row['assignment_group']).strip()
                            else:
                                assign_group = "Unknown"


                            def add_anomaly(issue_type, intella_st="", tms_st="", last_updated="", days_inactive=""):
                                record = {
                                    'Assignment Group': assign_group,
                                    'Responsible Person': assigned_to,
                                    'Issue Type': issue_type,
                                    'Case Number (JPR)': case_num,
                                    'TMS Number': tms_num,
                                    'Intella State': intella_st,
                                    'TMS Status': tms_st,
                                    'Last Updated Date': last_updated,
                                    'Days Inactive': days_inactive
                                }
                                if assign_group == 'RAS Agent':
                                    ras_anomalies.append(record)
                                else:
                                    non_ras_anomalies.append(record)


                            is_tms_missing = pd.isna(row['Requisition status'])
                            req_status = str(
                                row['Requisition status']).strip() if not is_tms_missing else "Missing / Not Found"

                            # === CHECK 0 & 1: Not Found OR Status Mismatch ===
                            if is_tms_missing:
                                not_in_tms_cases.append({
                                    'Assignment Group': assign_group,
                                    'Responsible Person': assigned_to,
                                    'Issue Type': 'Not Found in TMS',
                                    'Case Number (JPR)': case_num,
                                    'TMS Number': tms_num,
                                    'Intella State': state,
                                    'TMS Status': req_status,
                                    'Last Updated Date': '',
                                    'Days Inactive': ''
                                })
                            else:
                                # ✨ 匹配白名单逻辑
                                is_valid_match = (state == req_status) or (
                                            VALID_STATUS_MAPPING.get(state) == req_status)
                                if not is_valid_match:
                                    add_anomaly('Status Mismatch', state, req_status)

                            # === CHECK 2: GSSC Process Completed ===
                            if state == 'GSSC Process Completed' or req_status == 'GSSC Process Completed':
                                add_anomaly('Status is GSSC Process Completed', state, req_status)

                            # === CHECK 3: Offer issued AND > 60 days ===
                            if state == 'Offer issued' or req_status == 'Offer issued':
                                if pd.notna(row['sys_updated_on_dt']):
                                    days_diff = (today - row['sys_updated_on_dt']).days
                                    if days_diff > 60:
                                        formatted_date = row['sys_updated_on_dt'].strftime('%Y-%m-%d')
                                        add_anomaly('Offer Issued & Inactive > 60 Days', state, req_status,
                                                    formatted_date, days_diff)

                        st.markdown("---")

                        # =========================================================
                        # SECTION: NOT FOUND IN TMS (恢复)
                        # =========================================================
                        if not_in_tms_cases:
                            st.markdown("## 🛑 Intella IDs Not Found in TMS")
                            st.warning(
                                f"Found {len(not_in_tms_cases)} cases in Intella that do not exist in the combined TMS reports.")
                            df_not_in_tms = pd.DataFrame(not_in_tms_cases)
                            df_not_in_tms = df_not_in_tms.sort_values(by=['Assignment Group', 'Responsible Person'],
                                                                      ascending=[True, True]).reset_index(drop=True)

                            st.dataframe(df_not_in_tms, use_container_width=True)

                            output_missing = io.BytesIO()
                            with pd.ExcelWriter(output_missing, engine='openpyxl') as writer:
                                df_not_in_tms.to_excel(writer, sheet_name='Not Found in TMS', index=False)
                                worksheet_miss = writer.sheets['Not Found in TMS']
                                for idx, col in enumerate(df_not_in_tms.columns):
                                    max_data_len = df_not_in_tms[col].astype(str).map(
                                        len).max() if not df_not_in_tms.empty else 0
                                    col_len = max(max_data_len, len(str(col))) + 2
                                    worksheet_miss.column_dimensions[get_column_letter(idx + 1)].width = col_len

                            st.download_button(
                                label="📥 Download 'Not Found in TMS' Report (Excel)",
                                data=output_missing.getvalue(),
                                file_name='not_found_in_tms_report.xlsx',
                                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            )
                        else:
                            st.success("✅ All Intella IDs were successfully matched in TMS reports.")

                        st.markdown("<br><br>", unsafe_allow_html=True)

                        # =========================================================
                        # SECTION: RAS AGENT ANOMALIES
                        # =========================================================
                        if ras_anomalies:
                            st.markdown("## 🔹 Standard Processing (RAS Agent)")
                            df_anomalies = pd.DataFrame(ras_anomalies)
                            df_anomalies = df_anomalies.drop(columns=['Assignment Group'])
                            df_anomalies = df_anomalies.sort_values(by=['Responsible Person', 'Issue Type'],
                                                                    ascending=[True, True]).reset_index(drop=True)

                            # ✨ 注入空 Comment 列
                            df_anomalies['Comment'] = ""

                            # 恢复：生成 Summary 汇总表
                            summary_df = df_anomalies.groupby(['Responsible Person', 'Issue Type']).agg(
                                Case_Count=('Case Number (JPR)', 'count'),
                                Specific_JPRs=('Case Number (JPR)', lambda x: ', '.join(x.unique())),
                                Specific_TMS=('TMS Number', lambda x: ', '.join(x.unique()))
                            ).reset_index()

                            formatted_summary_data = []
                            unique_persons_sorted = sorted(df_anomalies['Responsible Person'].unique())

                            for person in unique_persons_sorted:
                                person_df = df_anomalies[df_anomalies['Responsible Person'] == person]
                                total_count = len(person_df)

                                formatted_summary_data.append({
                                    'Responsible Person': person,
                                    'Issue Type': 'Total Anomalies',
                                    'Case Count': total_count,
                                    'Specific JPRs': '',
                                    'Specific TMS Numbers': ''
                                })

                                person_summary = summary_df[summary_df['Responsible Person'] == person]
                                for _, r in person_summary.iterrows():
                                    formatted_summary_data.append({
                                        'Responsible Person': person,
                                        'Issue Type': f"  └ {r['Issue Type']}",
                                        'Case Count': r['Case_Count'],
                                        'Specific JPRs': r['Specific_JPRs'],
                                        'Specific TMS Numbers': r['Specific_TMS']
                                    })
                                formatted_summary_data.append(
                                    {'Responsible Person': '', 'Issue Type': '', 'Case Count': '', 'Specific JPRs': '',
                                     'Specific TMS Numbers': ''})

                            df_formatted_summary = pd.DataFrame(formatted_summary_data).iloc[:-1]
                            styled_summary = df_formatted_summary.style.apply(style_summary_cells, axis=1)

                            st.subheader("📊 Anomaly Summary")
                            st.dataframe(styled_summary, use_container_width=True)

                            output_summary = io.BytesIO()
                            with pd.ExcelWriter(output_summary, engine='openpyxl') as writer:
                                styled_summary.to_excel(writer, sheet_name='Anomaly Summary', index=False)
                                worksheet_sum = writer.sheets['Anomaly Summary']
                                for idx, col in enumerate(df_formatted_summary.columns):
                                    max_data_len = df_formatted_summary[col].astype(str).map(
                                        len).max() if not df_formatted_summary.empty else 0
                                    col_len = max(max_data_len, len(str(col))) + 2
                                    worksheet_sum.column_dimensions[get_column_letter(idx + 1)].width = col_len

                            st.download_button(
                                label="📥 Download Summary (Excel)",
                                data=output_summary.getvalue(),
                                file_name='ras_anomaly_summary.xlsx',
                                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            )

                            st.markdown("---")

                            # 带有 Comment 列的 Detailed List
                            st.subheader("🗂️ Detailed Anomaly List (Ready for SharePoint)")
                            styled_anomalies = df_anomalies.style.apply(color_cells, axis=1)
                            st.dataframe(styled_anomalies, use_container_width=True)

                            output_detailed = io.BytesIO()
                            with pd.ExcelWriter(output_detailed, engine='openpyxl') as writer:
                                for person in unique_persons_sorted:
                                    person_df = df_anomalies[df_anomalies['Responsible Person'] == person].copy()
                                    styled_person_df = person_df.style.apply(color_cells, axis=1)

                                    safe_sheet_name = re.sub(r'[\\/*?:\[\]]', '', str(person))
                                    safe_sheet_name = safe_sheet_name[:31] if safe_sheet_name else "Unknown"

                                    styled_person_df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
                                    worksheet_det = writer.sheets[safe_sheet_name]
                                    for idx, col in enumerate(person_df.columns):
                                        max_data_len = person_df[col].astype(str).map(
                                            len).max() if not person_df.empty else 0
                                        col_len = max(max_data_len, len(str(col))) + 2
                                        worksheet_det.column_dimensions[get_column_letter(idx + 1)].width = col_len

                            st.download_button(
                                label="📥 Download Detailed List (Excel)",
                                data=output_detailed.getvalue(),
                                file_name='ras_anomaly_detailed.xlsx',
                                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            )
                        else:
                            st.info("No RAS Agent anomalies found.")

                        st.markdown("<br><br>", unsafe_allow_html=True)

                        # =========================================================
                        # SECTION: NON-RAS AGENT ANOMALIES (恢复)
                        # =========================================================
                        if non_ras_anomalies:
                            st.markdown("## ⚠️ Requires Manual Screening (Non-RAS Agent)")
                            df_non_ras = pd.DataFrame(non_ras_anomalies)

                            df_non_ras = df_non_ras.sort_values(
                                by=['Assignment Group', 'Responsible Person', 'Issue Type'],
                                ascending=[True, True, True]
                            ).reset_index(drop=True)

                            styled_non_ras = df_non_ras.style.apply(color_cells, axis=1)
                            st.dataframe(styled_non_ras, use_container_width=True)

                            output_non_ras = io.BytesIO()
                            with pd.ExcelWriter(output_non_ras, engine='openpyxl') as writer:
                                styled_non_ras.to_excel(writer, sheet_name='Manual Screening', index=False)
                                worksheet_non_ras = writer.sheets['Manual Screening']
                                for idx, col in enumerate(df_non_ras.columns):
                                    max_data_len = df_non_ras[col].astype(str).map(
                                        len).max() if not df_non_ras.empty else 0
                                    col_len = max(max_data_len, len(str(col))) + 2
                                    worksheet_non_ras.column_dimensions[get_column_letter(idx + 1)].width = col_len

                            st.download_button(
                                label="📥 Download Manual Screening List (Excel)",
                                data=output_non_ras.getvalue(),
                                file_name='non_ras_manual_screening.xlsx',
                                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            )
                        elif ras_anomalies:
                            st.success("All found anomalies belong to RAS Agents. No manual screening required.")

        except Exception as e:
            st.error(f"An unexpected error occurred while processing the files: {e}")

# ---------------------------------------------------------
# TAB 2: TRACK PROGRESS (SharePoint File Review)
# ---------------------------------------------------------
with tab2:
    st.header("📈 Tracker: Review Progress & Pending Cases")
    st.write("Upload the updated 'Detailed Anomaly List' downloaded from SharePoint to check everyone's progress.")

    file_tracking = st.file_uploader("Upload Tracked Excel (.xlsx)", type=['xlsx'], key="tracker_upload")

    if file_tracking:
        with st.spinner('Calculating progress and extracting pending cases...'):
            try:
                xls = pd.ExcelFile(file_tracking)
                summary_data = []
                pending_cases = []

                for sheet_name in xls.sheet_names:
                    df_sheet = pd.read_excel(xls, sheet_name=sheet_name)

                    if 'Comment' in df_sheet.columns:
                        total_cases = len(df_sheet)
                        is_completed = df_sheet['Comment'].notna() & (df_sheet['Comment'].astype(str).str.strip() != '')
                        completed_cases = is_completed.sum()
                        progress = completed_cases / total_cases if total_cases > 0 else 0

                        summary_data.append({
                            'Responsible Person': sheet_name,
                            'Total Cases': total_cases,
                            'Completed': completed_cases,
                            'Pending': total_cases - completed_cases,
                            'Progress': progress
                        })

                        df_pending = df_sheet[~is_completed].copy()
                        if not df_pending.empty:
                            pending_cases.append(df_pending)

                if summary_data:
                    df_progress = pd.DataFrame(summary_data)
                    st.subheader("📊 Review Progress Summary")

                    # ✨ 视觉优化：为 Progress 列添加红-黄-绿的色阶渐变
                    styled_progress = df_progress.style.background_gradient(
                        subset=['Progress'],
                        cmap='RdYlGn',  # 颜色映射：Red -> Yellow -> Green
                        vmin=0.0,  # 最小值固定为 0 (0%)
                        vmax=1.0  # 最大值固定为 1 (100%)
                    ).format({'Progress': '{:.1%}'})

                    st.dataframe(styled_progress, use_container_width=True)

                if pending_cases:
                    df_all_pending = pd.concat(pending_cases, ignore_index=True)
                    st.subheader("⏳ All Pending Cases (Missing Comments)")
                    st.write(f"Found **{len(df_all_pending)}** pending cases across all agents.")
                    st.dataframe(df_all_pending, use_container_width=True)

                    output_pending = io.BytesIO()
                    with pd.ExcelWriter(output_pending, engine='openpyxl') as writer:
                        df_all_pending.to_excel(writer, index=False, sheet_name='Pending Cases')
                        worksheet_pending = writer.sheets['Pending Cases']
                        for idx, col in enumerate(df_all_pending.columns):
                            worksheet_pending.column_dimensions[get_column_letter(idx + 1)].width = 20

                    st.download_button(
                        label="📥 Download All Pending Cases (Excel)",
                        data=output_pending.getvalue(),
                        file_name="Pending_Cases_To_Chase.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                elif summary_data:
                    st.success("🎉 Amazing! All cases have been reviewed and commented by everyone.")

            except Exception as e:
                st.error(
                    f"Error processing the tracking file. Please ensure it is the correct Excel file. Error details: {e}")

st.markdown("---")
st.markdown(
    "<div style='text-align: center; color: gray; font-size: 14px; margin-top: 1px;'>&copy; 2026 Li Zihan. All rights reserved. <br><span style='font-size: 12px;'>RAS Monitoring Report Analysis Dashboard v2.0</span></div>",
    unsafe_allow_html=True)
